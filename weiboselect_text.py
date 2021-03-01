import pandas as pd
from pandas import Series, DataFrame # 主要数据结构
import openpyxl
import re

w = openpyxl.load_workbook(filename=r'D:\CS_zjw\Weibo_projects\test0105\comd\comd_classify_text.xlsx')
openpyxl.sheet_ranges = w['Sheet1']
ws = w.active

wb_ctg = pd.read_excel(r'D:\CS_zjw\Weibo_projects\test0105\words_weibo_emoji0.xlsx')
data_wb = pd.read_excel(r'D:\CS_zjw\Weibo_projects\test0105\comd\allsample_commend.xlsx')
data_wbtxt=pd.DataFrame(data_wb['content'])
index=len(list(data_wbtxt.index))
#print(data_wbtxt[83])
#print(len(list(data_wbtxt.index)))
cnt=0
for i in range(index):
    f=dict(data_wbtxt.loc[i])['content']
    emoji_wrds=re.findall(r'\[([^\[\]]+)\]',f)
    ws['A'+str(cnt+1)] = cnt
    ws['B'+str(cnt+1)] = str(f)
    ws['C'+str(cnt+1)] = str(emoji_wrds)
    #print(str(emoji_wrds)) #emoji_wrds的类型是list
    cnt+=1
#print(wb_ctg)
w.save(r'D:\CS_zjw\Weibo_projects\test0105\comd\comd_classify_text.xlsx')
