import pandas as pd
from pandas import Series, DataFrame # 主要数据结构
import openpyxl
import re
import emoji
'''
w = openpyxl.load_workbook(filename=r'D:\CS_zjw\Weibo_projects\test0105\comd\comd_classify_figure.xlsx')
openpyxl.sheet_ranges = w['Sheet1']
ws = w.active
'''
figure_ctg = pd.read_excel(r'D:\CS_zjw\Weibo_projects\test0105\words_figure_emoji0.xlsx')
data_wb = pd.read_excel(r'D:\CS_zjw\Weibo_projects\test0105\comd\allsample_commend.xlsx')
data_wbtxt=pd.DataFrame(data_wb['content'])
index=len(list(data_wbtxt.index))
#print(data_wbtxt[83])
#print(len(list(data_wbtxt.index)))
figure_po=list(figure_ctg['positive'][:30])
figure_neg=list(figure_ctg['negative'][:44])
figure_neu=list(figure_ctg['neutral'][:24])

print(figure_po)
print(figure_neg)
print(figure_neu)
'''for p in range(len(figure_po)):
    q=emoji.demojize(figure_po[p])
    ww=emoji.emojize(q)
    emoji_cnt=emoji.emoji_count(figure_po[p])
    print(p,',',c)
    '''
'''
cnt=0
cntexl=cnt+2
for i in range(index):
    cntexl=cnt+2
    f=dict(data_wbtxt.loc[i])['content']
    ws['A'+str(cntexl)] = cnt
    ws['B'+str(cntexl)] = f
    for j in range(len(f)):
        obj=f[j]
        emoji_cnt=emoji.emoji_count(f)
        ws['E'+str(cntexl)] = emoji_cnt
        
        if obj in figure_po:
            #print('po'+obj)
            ws['C'+str(cntexl)] = obj
            ws['D'+str(cntexl)] = 'positive'
        elif obj in figure_neg:
            #print('neg'+obj)
            ws['C'+str(cntexl)] = obj
            ws['D'+str(cntexl)] = 'negative'
        elif obj in figure_neu:
            #print('neu'+obj)
            ws['C'+str(cntexl)] = obj
            ws['D'+str(cntexl)] = 'neutral'

        #print(str(emoji_wrds)) #emoji_wrds的类型是list
    cnt+=1
#print(wb_ctg)
w.save(r'D:\CS_zjw\Weibo_projects\test0105\comd\comd_classify_figure.xlsx')
'''
